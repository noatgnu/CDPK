import re
import csv
import os
from sequon import settingsequon
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from tornado import gen

def file_set(input):
    input_path = os.path.join(settingsequon.UPLOAD_FOLDER, input)
    filen = input.replace('.txt', '').replace('.csv', '')
    return input_path, filen

@gen.coroutine
def file_opener_s(input_path, filen):
    response = yield file_opener(input_path, filen)
    return response.body

def header_generation(input_path):
    fieldnames = []
    with open(input_path, 'rt', newline='') as csvfile:
        reader = csv.reader(csvfile, dialect = 'excel', delimiter='\t')
        row = next(reader)
        if not len(fieldnames) == len(row):
                for k in row:
                    if not k in fieldnames:
                        fieldnames.append(k)
    return fieldnames

def file_opener(input_path, filen, fieldnames):
    nxst_all = os.path.join(settingsequon.RESULTS_FOLDER, filen+'_NXST_all.csv')
    nxs_nohexnac = os.path.join(settingsequon.RESULTS_FOLDER, filen+'_NXS_noHexNAc.csv')
    nxt_nohexnac = os.path.join(settingsequon.RESULTS_FOLDER, filen+'_NXT_noHexNAc.csv')
    nxs_hexnac = os.path.join(settingsequon.RESULTS_FOLDER, filen+'_NXS_HexNAc.csv')
    nxt_hexnac = os.path.join(settingsequon.RESULTS_FOLDER, filen+'_NXT_HexNAc.csv')
    with open(input_path, 'rt', newline='') as csvfile, open(nxst_all, 'wt', newline='', encoding='utf-8' ) as outputcsvfilenxst_all, open(nxs_nohexnac, 'wt', newline='', encoding='utf-8' ) as outputcsvfilenxs_nohexnac, open(nxt_nohexnac, 'wt', newline='', encoding='utf-8') as outputcsvfilenxt_nohexnac, open(nxs_hexnac, 'wt', newline='', encoding='utf-8') as outputcsvfilenxs_hexnac, open(nxt_hexnac, 'wt', newline='', encoding='utf-8') as outputcsvfilenxt_hexnac:        
        summarypeptide = csv.DictReader(csvfile, dialect='excel', delimiter='\t')
        writenxst_all = csv.DictWriter(outputcsvfilenxst_all, fieldnames=fieldnames, dialect='excel')
        writenxst_all.writeheader()
        writenxs_hexnac = csv.DictWriter(outputcsvfilenxs_hexnac, fieldnames=fieldnames, dialect='excel')
        writenxs_hexnac.writeheader()
        writenxt_hexnac = csv.DictWriter(outputcsvfilenxt_hexnac, fieldnames=fieldnames, dialect='excel')
        writenxt_hexnac.writeheader()
        writenxs_nohexnac = csv.DictWriter(outputcsvfilenxs_nohexnac, fieldnames=fieldnames, dialect='excel')
        writenxs_nohexnac.writeheader()
        writenxt_nohexnac = csv.DictWriter(outputcsvfilenxt_nohexnac, fieldnames=fieldnames, dialect='excel')
        writenxt_nohexnac.writeheader()
        for row in summarypeptide:
            sequon_search = re.search('(N\w(S|T))', row['Sequence'])
            
            if sequon_search:
                if not sequon_search.group(0)[1] == 'P':
                    writenxst_all.writerow(row)
                    modification_search = re.search('(HexNAc)', row['ProteinModifications'])
                    if sequon_search.group(0)[2] == 'S':
                        if modification_search:
                            writenxs_hexnac.writerow(row)
                        else:
                            writenxs_nohexnac.writerow(row)
                    if sequon_search.group(0)[2] == 'T':
                        if modification_search:
                            writenxt_hexnac.writerow(row)
                        else:
                            writenxt_nohexnac.writerow(row)
    return {filen: [filen+'_NXST_all.csv', filen+'_NXS_HexNAc.csv', filen+'_NXS_noHexNAc.csv', filen+'_NXT_HexNAc.csv', filen+'_NXT_noHexNAc.csv']}

def csv_to_excel_converter(csvlistresult):
    excel_output_path = {}
    excel_dict = {}
    sname_list = []    
    for k in csvlistresult:
        input_list = []
        for i in csvlistresult[k]:
            input_list.append(os.path.join(settingsequon.RESULTS_FOLDER, i))
            sname = re.search('(NX\w*).csv$', i)
            sname_list.append(sname.group(1))
        excel_dict[k] = input_list
    for k in excel_dict:
        wb = Workbook()
        for i in sname_list:
            ws = wb.create_sheet()
            ws.title = i
        for i in excel_dict[k]:
            d = re.search('(NX\w*).csv$', i)
            if d.group(1) in sname_list:
                with open(i, 'rt', newline='') as nxf:
                    reader = csv.reader(nxf, dialect = 'excel')                
                    ws = wb.get_sheet_by_name(d.group(1))
                    for row_index, row in enumerate(reader):
                        for column_index, cell in enumerate(row):
                            column_letter = get_column_letter((column_index + 1))
                            s = cell
                            try:
                                s=float(s)
                            except ValueError:
                                pass

                            ws.cell('%s%s' % (column_letter, (row_index + 1))).value = s
        ws = wb.get_sheet_by_name('Sheet')
        wb.remove_sheet(ws)
        wb.save(filename = os.path.join(settingsequon.RESULTS_FOLDER, k+"_NXST_all.xlsx"))
        excel_output_path[k] = k+"_NXST_all.xlsx"
    return excel_output_path
                
        
def allowed_file(filename):
    split_fn = filename.split('.')[-1]
    if split_fn:
       if split_fn in settingsequon.ALLOWED_EXTENSIONS:
           return 'True'
       else:
           return 'False'
    return 'False'

def processor(filen):
    input_fp, fn = file_set(filen)
    fieldnames = header_generation(input_fp)
    result_fn = file_opener(input_fp, fn, fieldnames)
    excel = csv_to_excel_converter(result_fn)
    return result_fn, excel